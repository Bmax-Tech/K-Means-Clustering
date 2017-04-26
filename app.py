from flask import Flask, jsonify, request
from openpyxl import load_workbook
import algorithm
import numpy as np

app = Flask(__name__)


# ----------------------------------------------------------------------------------------------------------------------
# Custom Functions
# ----------------------------------------------------------------------------------------------------------------------

def load_data_set():
    # main array
    data_array = []
    # load xlsx data set
    wb = load_workbook('data/data_set_2.xlsx')
    print(wb.sheetnames)
    for val in wb.sheetnames:
        ws = wb.get_sheet_by_name(val)
        heading = False
        headings = []
        for row in ws.iter_rows():
            # local array
            temp_1 = []
            for cell in row:
                cell_content = str(cell.value)
                temp_1.append(cell_content)
            if not heading:
                headings = temp_1
            else:
                # set key value pairs
                data = {}
                for x in range(len(headings)):
                    data[headings[x]] = temp_1[x]
                data_array.append(data)
            heading = True
    return data_array


# ----------------------------------------------------------------------------------------------------------------------
# API - Routing
# ----------------------------------------------------------------------------------------------------------------------

@app.route("/getData", methods=['POST'])
def index():
    data_points, sample_preds = algorithm.cluster_results(request.json)
    res = {}
    res['sample_preds'] = []
    res['data_points'] = []
    for i, point in enumerate(data_points):
        temp = {}
        temp['x'] = point[0]
        temp['y'] = point[1]
        res['data_points'].append(temp)

    for i, pred in enumerate(sample_preds):
        res['sample_preds'].append(np.int32(pred).item())

    return jsonify(res)


if __name__ == "__main__":
    app.run(
        host="127.0.0.1",
        port=int(7890)
    )
